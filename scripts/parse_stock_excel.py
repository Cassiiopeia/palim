#!/usr/bin/env python3
"""엑셀/CSV 를 JSON 으로 변환한다.

호출 규약(04-CONVENTIONS):
  - stdout 에는 JSON 만. 사람용 메시지는 stderr.
  - 종료코드 0(성공) / 1(실패).
  - 인자는 배열로 받는다. 쉘 문자열을 조립하지 않는다.

사용:
  python3 parse_stock_excel.py <파일경로> [--header-row N] [--limit N] [--sheet NAME]

--limit 은 미리보기용이다. 지정해도 row_count 는 전체 건수를 보고한다 —
화면이 "N건 중 5건 미리보기"를 표시해야 하기 때문이다.

pandas 를 쓰지 않는다. 필요한 기능이 "헤더를 읽고 각 행을 문자열 dict 로" 뿐인데
pandas 는 수십 MB 이고, 값 타입을 추론했다가 다시 문자열로 되돌리는 낭비가 생긴다.
CSV 는 표준 csv 모듈로 처리하고, 엑셀일 때만 openpyxl 을 지연 import 한다 —
CSV 만 다루는 환경에서는 외부 의존이 아예 필요 없다.
"""

import argparse
import csv
import datetime
import json
import os
import sys

EXCEL_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")


def normalize_header(value) -> str:
    """헤더 정리. 원본에 줄바꿈·앞뒤 공백이 섞여 있는 경우가 흔하다."""
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ").replace("\r", " ")


def to_text(value) -> str:
    """셀 값을 문자열로.

    엑셀은 수량 120 을 120.0(float) 으로 준다. 그대로 두면 매핑 단계에서
    "120.0" 을 정수로 읽어야 하는 지저분한 처리가 따라붙는다. 정수값 float 은
    소수점을 떼고, 날짜는 ISO 문자열로 통일한다.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return str(value).strip()


def read_csv_file(path: str, header_index: int):
    """CSV 읽기. BOM 이 붙은 파일이 흔해 utf-8-sig 로 연다."""
    with open(path, newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))

    if header_index >= len(rows):
        raise ValueError(f"헤더 행({header_index + 1})이 파일 범위를 벗어납니다")

    headers = [normalize_header(cell) for cell in rows[header_index]]
    body = [[to_text(cell) for cell in row] for row in rows[header_index + 1:]]
    return headers, body


def read_excel_file(path: str, header_index: int, sheet: str | None):
    """엑셀 읽기. read_only 로 열어 큰 파일에서 메모리를 아낀다."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            "엑셀을 읽으려면 openpyxl 이 필요합니다. pip install openpyxl") from None

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.worksheets[0]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if header_index >= len(rows):
        raise ValueError(f"헤더 행({header_index + 1})이 파일 범위를 벗어납니다")

    headers = [normalize_header(cell) for cell in rows[header_index]]
    body = [[to_text(cell) for cell in row] for row in rows[header_index + 1:]]
    return headers, body


def build_rows(headers: list[str], body: list[list[str]]) -> list[dict]:
    """헤더 이름이 있는 열만 남긴다. 빈 헤더는 서식용 빈 열이다."""
    keep = [index for index, name in enumerate(headers) if name]

    result = []
    for row in body:
        # 전부 빈 행은 파일 끝 여백이다. 넣으면 "필수 항목 없음"으로 무더기 실패한다.
        if not any(str(cell).strip() for cell in row):
            continue
        record = {}
        for index in keep:
            record[headers[index]] = row[index] if index < len(row) else ""
        result.append(record)
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--header-row", type=int, default=1,
                        help="헤더가 있는 행 번호(1부터). 기본 1")
    parser.add_argument("--limit", type=int, default=0,
                        help="0이면 전체. 미리보기는 5 정도")
    parser.add_argument("--sheet", default=None)
    args = parser.parse_args()

    if not os.path.isfile(args.path):
        print(f"파일을 찾을 수 없습니다: {args.path}", file=sys.stderr)
        return 1

    header_index = max(args.header_row - 1, 0)

    try:
        if args.path.lower().endswith(EXCEL_SUFFIXES):
            headers, body = read_excel_file(args.path, header_index, args.sheet)
        else:
            headers, body = read_csv_file(args.path, header_index)
    except (ValueError, RuntimeError) as exc:
        print(str(exc), file=sys.stderr)
        return 1
    except Exception as exc:  # 파일 손상·권한 등 종류가 많아 하나로 묶는다
        print(f"파일을 읽을 수 없습니다: {exc}", file=sys.stderr)
        return 1

    rows = build_rows(headers, body)
    total = len(rows)
    if args.limit > 0:
        rows = rows[:args.limit]

    json.dump(
        {"fields": [name for name in headers if name], "rows": rows, "row_count": total},
        sys.stdout, ensure_ascii=False,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
