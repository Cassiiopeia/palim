#!/usr/bin/env python3
"""대조 결과를 엑셀 파일로 쓴다.

호출 규약(04-CONVENTIONS):
  - stdout 에는 JSON 만. 사람용 메시지는 stderr.
  - 종료코드 0(성공) / 1(실패).
  - 인자는 배열로 받는다. 쉘 문자열을 조립하지 않는다.

사용:
  python3 write_reconcile_xlsx.py <입력JSON경로> <출력xlsx경로>

입력을 인자가 아니라 «파일» 로 받는 이유 — 결과가 수천 줄이 될 수 있고, 명령행 길이 상한을
넘으면 그때부터 조용히 잘린다. 잘린 파일은 «틀린 파일» 이 아니라 «짧은 파일» 이라 사람이
알아채지 못한다.

## 품목코드를 글자로 고정하는 이유

엑셀은 숫자처럼 «생긴» 값을 숫자로 해석한다. 그러면 품목코드 00094 가 94 로, 01002 가
1,002 로 바뀐다. 이 제품의 핵심이 그 코드를 견주는 것이므로, 내려받은 파일에서 코드가
달라지면 그 파일로는 아무것도 못 맞춘다.

openpyxl 에서 셀 서식을 '@'(텍스트)로 두면 엑셀이 값을 그대로 보여준다.
"""

import json
import sys


def fail(message):
    print(message, file=sys.stderr)
    sys.exit(1)


def write_sheet(workbook, title, columns, rows, first_index):
    """한 장을 쓴다. first_index 앞쪽 칸은 글자로 고정한다(코드·이름)."""
    sheet = workbook.create_sheet(title=title)
    sheet.append(columns)

    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    # 글자로 고정. 서식만 바꾸면 openpyxl 이 이미 숫자로 넣은 값은 되돌아오지 않으므로,
    # 애초에 문자열로 넣고(위 append 가 그렇다) 서식으로 «엑셀이 다시 해석하는 것» 을 막는다.
    for line in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(line):
            if index < first_index:
                cell.number_format = "@"

    # 칸 너비. 좁으면 ### 로 보이고, 그러면 값을 읽으려고 매번 늘려야 한다.
    for index, column in enumerate(columns, start=1):
        longest = max([len(str(column))] + [len(str(r.get(column, ""))) for r in rows] or [0])
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "AA"].width = \
            min(max(longest + 2, 10), 40)

    sheet.freeze_panes = "A2"
    return sheet


def main():
    if len(sys.argv) != 3:
        fail("사용: write_reconcile_xlsx.py <입력JSON경로> <출력xlsx경로>")

    source, target = sys.argv[1], sys.argv[2]

    try:
        with open(source, encoding="utf-8") as handle:
            payload = json.load(handle)
    except (OSError, ValueError) as error:
        fail(f"입력을 읽지 못했습니다: {error}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        fail("openpyxl 이 없습니다. python3-openpyxl 을 설치하세요.")

    global HEADER_FONT, HEADER_ALIGN
    HEADER_FONT = Font(bold=True)
    HEADER_ALIGN = Alignment(vertical="center")

    workbook = Workbook()
    # 기본으로 생기는 빈 장을 지운다. 남겨 두면 파일을 열었을 때 빈 화면이 먼저 보인다.
    workbook.remove(workbook.active)

    # 첫 장은 «무엇을 견줬나» 다. 며칠 뒤에 열면 화면과 짝지을 수 없으므로 파일이 스스로
    # 말해야 한다.
    summary = workbook.create_sheet(title="이 대조는")
    for line in payload.get("summary", []):
        summary.append([line.get("label", ""), line.get("value", "")])
    for row in summary.iter_rows(min_col=1, max_col=1):
        for cell in row:
            cell.font = HEADER_FONT
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 60

    total = 0
    for sheet_spec in payload.get("sheets", []):
        rows = sheet_spec.get("rows", [])
        total += len(rows)
        write_sheet(workbook, sheet_spec.get("title", "결과"),
                    sheet_spec.get("columns", []), rows,
                    sheet_spec.get("textColumns", 0))

    try:
        workbook.save(target)
    except OSError as error:
        fail(f"파일을 쓰지 못했습니다: {error}")

    json.dump({"path": target, "rows": total}, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
